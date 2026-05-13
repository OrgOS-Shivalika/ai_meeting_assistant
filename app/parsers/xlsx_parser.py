"""XLSX parser — openpyxl, one block per row.

Spreadsheets defeat naive text extraction: pasting a sheet as raw CSV
loses column context, and stuffing the whole grid into one block
shreds retrieval ranking. Strategy:

  - One ParsedBlock per non-empty row.
  - Block text is rendered as `"<header1>: <value1> | <header2>: <value2>"`
    when the sheet has an identifiable header row (the first non-empty
    row), or as `"<col_A>: <value> | <col_B>: <value>"` otherwise.
    Header-attaching is the trick that makes XLSX rows retrievable —
    "what was Q3 revenue for the EU team?" never matches a bare number.
  - `section_path` is the sheet name.
  - `metadata.row_number` carries the 1-indexed openpyxl row for
    citations ("sheet 'Q3 forecast' row 14").

We use `read_only=True` + `data_only=True` so formulas surface as their
last-cached value (matches what a human reading the file sees).
"""
from __future__ import annotations

import io
import logging
import re

from app.parsers.base import ParsedBlock

logger = logging.getLogger(__name__)

_WS_RX = re.compile(r"\s+")


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # Drop trailing zeros for nicer embeddings: 1.0 -> "1", 1.50 -> "1.5"
        if v.is_integer():
            return str(int(v))
        return ("%g" % v).strip()
    return str(v).strip()


def _is_header_row(values: list[str]) -> bool:
    """A header row is one where every non-empty cell is a string and at
    least one is alphabetic — i.e. no all-numeric rows."""
    nonempty = [v for v in values if v]
    if not nonempty:
        return False
    has_letter = False
    for v in nonempty:
        if any(c.isalpha() for c in v):
            has_letter = True
        # If any "header" looks numeric, treat the whole row as data.
        try:
            float(v.replace(",", ""))
        except ValueError:
            continue
        else:
            return False
    return has_letter


def parse_xlsx(raw_bytes: bytes) -> list[ParsedBlock]:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except InvalidFileException as e:
        logger.warning("openpyxl rejected file (%s)", e)
        return []
    except Exception as e:  # noqa: BLE001
        logger.warning("openpyxl unexpected error (%s)", e)
        return []

    blocks: list[ParsedBlock] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        headers: list[str] | None = None
        first_row_idx_seen = 0

        for row_idx, row in enumerate(rows_iter, start=1):
            values = [_cell_str(v) for v in row]
            # Empty rows: skip but keep counting so row_number stays
            # aligned with what the user sees in Excel.
            if not any(values):
                continue

            if headers is None:
                first_row_idx_seen = row_idx
                if _is_header_row(values):
                    # Backfill blanks so column index stays meaningful.
                    headers = [v if v else f"col_{chr(ord('A') + i)}"
                               for i, v in enumerate(values)]
                    continue
                # No header — synthesize column-letter headers so every
                # cell still gets a label.
                headers = [f"col_{chr(ord('A') + i)}" for i in range(len(values))]

            # Render this row as "<header>: <value>" pairs, skipping
            # empty cells so the embedding isn't padded with " | : | : |".
            parts: list[str] = []
            for h, v in zip(headers, values):
                v_clean = _WS_RX.sub(" ", v).strip()
                if not v_clean:
                    continue
                parts.append(f"{h}: {v_clean}")
            if not parts:
                continue

            blocks.append(
                ParsedBlock(
                    block_index=len(blocks),
                    text=" | ".join(parts),
                    page_number=None,
                    section_path=sheet_name,
                    metadata={
                        "source_subtype": "xlsx",
                        "sheet_name": sheet_name,
                        "row_number": row_idx,
                        # Lets the chunker stitch the header into the
                        # first chunk of each sheet for context.
                        "first_data_row": (row_idx == first_row_idx_seen + 1
                                           if headers else False),
                    },
                )
            )
    return blocks
